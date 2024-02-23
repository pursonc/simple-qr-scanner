from flask import Flask, request, render_template
import os
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/qr_data', methods=['POST'])
def qr_data():
    data = request.form['qr_data']
    print("QR Code Data:", data)

    append_to_excel(data)
    # hadling data here
    return 'Received'

def append_to_excel(data):
    # load Excel file
    try:
        wb = load_workbook("qr_data.xlsx")
    except FileNotFoundError:
        # file not existed, craete Excel workbook
        wb = Workbook()
        
    # select the workbook
    ws = wb.active
    # append to workbook
    ws.append([data])
    # save the workbook
    wb.save("signin.xlsx")

if __name__ == '__main__':
    certfile = os.path.join(os.path.dirname(__file__), 'certs/cert.pem')
    keyfile = os.path.join(os.path.dirname(__file__), 'certs/key.pem')
    app.run(host='0.0.0.0', port=16888, ssl_context=(certfile, keyfile), debug=False)


